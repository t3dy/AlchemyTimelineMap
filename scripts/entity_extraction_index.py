#!/usr/bin/env python3
r"""
Entity Extraction Index Generator — ALCHEMYTIMELINEMAP

Reads all working notes from staging/working_notes/ and generates a comprehensive
Excel spreadsheet tracking every extracted entity (persons, texts, locations, concepts).

Used for: batch processing, deduplication, priority assignment, status tracking.
"""

import json
import re
import sys
import io
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set

# Force UTF-8 output on Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARNING] openpyxl not installed. Install with: pip install openpyxl")

# Configuration
WORKING_NOTES_DIR = Path("C:\\Dev\\ALCHEMYTIMELINEMAP\\staging\\working_notes")
OUTPUT_FILE = Path("C:\\Dev\\ALCHEMYTIMELINEMAP\\staging\\entity_extraction_index.xlsx")
SCHEMA_FILE = Path("C:\\Dev\\ALCHEMYTIMELINEMAP\\SCHEMA.json")

def slugify(name: str) -> str:
    """Convert name to slug format (lowercase, hyphens)."""
    return re.sub(r'[^\w\s-]', '', name.lower()).replace(' ', '-')

def load_schema() -> Dict:
    """Load SCHEMA.json to check existing entities."""
    if SCHEMA_FILE.exists():
        with open(SCHEMA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def extract_entities_from_working_note(note_path: Path) -> Dict[str, List[Dict]]:
    """Parse a working note markdown file and extract entities with metadata."""

    with open(note_path, 'r', encoding='utf-8') as f:
        content = f.read()

    entities = {
        "persons": [],
        "texts": [],
        "locations": [],
        "concepts": [],
    }

    source_file = re.search(r"\*\*Source:\*\*\s*(.+?)(?:\n|$)", content)
    source_name = source_file.group(1).strip() if source_file else "UNKNOWN"

    # Extract persons section
    persons_section = re.search(
        r"## Persons Identified\n\n(.*?)(?:\n\n### Notes on Persons|\n\n## )",
        content,
        re.DOTALL
    )
    if persons_section:
        lines = persons_section.group(1).split('\n')
        for line in lines:
            if line.startswith('- **'):
                name_match = re.search(r'- \*\*(.+?)\*\*:', line)
                if name_match:
                    name = name_match.group(1).strip()
                    status = "NEW" if "NEW" in line else "EXISTS"
                    entities["persons"].append({
                        "name": name,
                        "slug": slugify(name),
                        "status": status,
                        "source": source_name,
                        "source_file": note_path.stem,
                        "confidence": "MEDIUM"
                    })

    # Extract texts section
    texts_section = re.search(
        r"## Texts Identified\n\n(.*?)(?:\n\n### Notes on Texts|\n\n## )",
        content,
        re.DOTALL
    )
    if texts_section:
        lines = texts_section.group(1).split('\n')
        for line in lines:
            if line.startswith('- **'):
                title_match = re.search(r'- \*\*(.+?)\*\*:', line)
                if title_match:
                    title = title_match.group(1).strip()
                    status = "NEW" if "NEW" in line else "EXISTS"
                    entities["texts"].append({
                        "title": title,
                        "slug": slugify(title),
                        "status": status,
                        "source": source_name,
                        "source_file": note_path.stem,
                        "confidence": "MEDIUM"
                    })

    # Extract locations section
    locations_section = re.search(
        r"## Locations & Geographic Clusters\n\n(.*?)(?:\n\n## )",
        content,
        re.DOTALL
    )
    if locations_section:
        lines = locations_section.group(1).split('\n')
        for line in lines:
            if line.startswith('- **'):
                location_match = re.search(r'- \*\*(.+?)\*\*:', line)
                if location_match:
                    location = location_match.group(1).strip()
                    entities["locations"].append({
                        "name": location,
                        "slug": slugify(location),
                        "status": "EXISTS",  # Locations often already exist
                        "source": source_name,
                        "source_file": note_path.stem,
                        "confidence": "HIGH"
                    })

    # Extract concepts section
    concepts_section = re.search(
        r"## Concepts & Operations\n\n(.*?)(?:\n\n## )",
        content,
        re.DOTALL
    )
    if concepts_section:
        lines = concepts_section.group(1).split('\n')
        for line in lines:
            if line.startswith('- **'):
                concept_match = re.search(r'- \*\*(.+?)\*\*:', line)
                if concept_match:
                    concept = concept_match.group(1).strip()
                    status = "NEW" if "NEW" in line else "EXISTS"
                    entities["concepts"].append({
                        "term": concept,
                        "slug": slugify(concept),
                        "status": status,
                        "source": source_name,
                        "source_file": note_path.stem,
                        "confidence": "MEDIUM"
                    })

    return entities

def generate_excel_index(working_notes: List[Path], schema: Dict) -> None:
    """Generate Excel workbook with entity index."""

    if not HAS_OPENPYXL:
        print("Cannot generate Excel without openpyxl. Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Collect all entities
    all_persons = {}
    all_texts = {}
    all_locations = {}
    all_concepts = {}

    print(f"Parsing {len(working_notes)} working notes...")

    for note_path in working_notes:
        entities = extract_entities_from_working_note(note_path)

        for person in entities["persons"]:
            slug = person["slug"]
            if slug not in all_persons:
                all_persons[slug] = person
            else:
                # Merge if seen multiple times
                all_persons[slug]["source_file"] += f", {person['source_file']}"

        for text in entities["texts"]:
            slug = text["slug"]
            if slug not in all_texts:
                all_texts[slug] = text
            else:
                all_texts[slug]["source_file"] += f", {text['source_file']}"

        for location in entities["locations"]:
            slug = location["slug"]
            if slug not in all_locations:
                all_locations[slug] = location
            else:
                all_locations[slug]["source_file"] += f", {location['source_file']}"

        for concept in entities["concepts"]:
            slug = concept["slug"]
            if slug not in all_concepts:
                all_concepts[slug] = concept
            else:
                all_concepts[slug]["source_file"] += f", {concept['source_file']}"

    # Create sheets
    _create_entity_sheet(wb, "Persons", all_persons.values())
    _create_entity_sheet(wb, "Texts", all_texts.values())
    _create_entity_sheet(wb, "Locations", all_locations.values())
    _create_entity_sheet(wb, "Concepts", all_concepts.values())

    # Create summary sheet
    _create_summary_sheet(
        wb,
        len(all_persons),
        len(all_texts),
        len(all_locations),
        len(all_concepts),
        len(working_notes)
    )

    # Save
    wb.save(OUTPUT_FILE)
    print(f"[OK] Entity index saved: {OUTPUT_FILE}")
    print(f"   Persons: {len(all_persons)}")
    print(f"   Texts: {len(all_texts)}")
    print(f"   Locations: {len(all_locations)}")
    print(f"   Concepts: {len(all_concepts)}")

def _create_entity_sheet(wb, sheet_name: str, entities) -> None:
    """Create a sheet for one entity type."""

    ws = wb.create_sheet(sheet_name)

    # Header row
    if sheet_name == "Persons":
        headers = ["Slug", "Name", "Status", "Confidence", "Source File", "Source"]
    elif sheet_name == "Texts":
        headers = ["Slug", "Title", "Status", "Confidence", "Source File", "Source"]
    elif sheet_name == "Locations":
        headers = ["Slug", "Name", "Status", "Confidence", "Source File", "Source"]
    elif sheet_name == "Concepts":
        headers = ["Slug", "Term", "Status", "Confidence", "Source File", "Source"]

    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for entity in sorted(entities, key=lambda e: e.get("slug", "")):
        if sheet_name == "Persons":
            name_key = "name"
        elif sheet_name == "Texts":
            name_key = "title"
        elif sheet_name == "Locations":
            name_key = "name"
        elif sheet_name == "Concepts":
            name_key = "term"

        row = [
            entity.get("slug", ""),
            entity.get(name_key, ""),
            entity.get("status", "NEW"),
            entity.get("confidence", "MEDIUM"),
            entity.get("source_file", ""),
            entity.get("source", ""),
        ]
        ws.append(row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20

    # Add filter
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

def _create_summary_sheet(wb, num_persons: int, num_texts: int, num_locations: int, num_concepts: int, num_sources: int) -> None:
    """Create summary sheet."""

    ws = wb.create_sheet("Summary", 0)

    ws['A1'] = "Entity Extraction Index"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().isoformat()

    ws['A5'] = "Entity Type"
    ws['B5'] = "Count"
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)

    ws['A6'] = "Persons"
    ws['B6'] = num_persons
    ws['A7'] = "Texts"
    ws['B7'] = num_texts
    ws['A8'] = "Locations"
    ws['B8'] = num_locations
    ws['A9'] = "Concepts"
    ws['B9'] = num_concepts

    ws['A11'] = "Sources Processed"
    ws['B11'] = num_sources
    ws['A11'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

if __name__ == "__main__":
    print("Entity Extraction Index Generator")
    print("=" * 50)

    if not WORKING_NOTES_DIR.exists():
        print(f"❌ Working notes directory not found: {WORKING_NOTES_DIR}")
        print("   Run: python scripts/read_markdown_batch.py [limit]")
        exit(1)

    # Find all working notes
    working_notes = sorted(WORKING_NOTES_DIR.glob("note_*.md"))

    if not working_notes:
        print(f"❌ No working notes found in {WORKING_NOTES_DIR}")
        exit(1)

    # Load schema
    schema = load_schema()

    # Generate index
    generate_excel_index(working_notes, schema)

    print("\n[NEXT] Steps:")
    print(f"   1. Open {OUTPUT_FILE} in Excel")
    print("   2. Sort/filter by Status = 'NEW' to prioritize new entities")
    print("   3. Group by Source File to coordinate enrichment batches")
    print("   4. Assign confidence levels and review_status flags")
